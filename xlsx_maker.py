# Column filters
# https://openpyxl.readthedocs.io/en/stable/filters.html#using-filters-and-sorts



import os, datetime as dt
from openpyxl import Workbook
from openpyxl.worksheet.filters import (
    FilterColumn,
    CustomFilter,
    CustomFilters,
    DateGroupItem,
    Filters,
)
from openpyxl.styles import PatternFill



def Setup():
    wb = Workbook()
    wb.remove(wb.active)
    wb.create_sheet('Analytics')
    wb.create_sheet('Findings')
    ws_analytics = wb["Analytics"]
    ws_findings = wb["Findings"]
    return wb, ws_analytics, ws_findings


def SaveFile(wb, name):
    # YYYYMMDD
    timestamp_today = str(dt.date.today()).replace('-', '')

    reports_path = './_reports/' + timestamp_today

    if not os.path.exists(reports_path):
        os.makedirs(reports_path)

    wb.save(reports_path + '/' + name + '.xlsx')


def SetColumnColors(ws, column_colors):
    for idx, row in enumerate(ws.rows):
        for col in row:
            ws[col.coordinate].fill = PatternFill(start_color=column_colors[idx], end_color=column_colors[idx], fill_type='solid')


def SetColumnSize(ws):
    for col in ws.columns:
        max_length = 6
        column_letter = col[0].column_letter

        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass

        adjusted_width = (max_length + 2)
        ws.column_dimensions[column_letter].width = adjusted_width


def SetFindingsHyperlinks(ws_findings):
    for col in ws_findings.columns:
        column_letter = col[0].column_letter

        # hard coding column with hyperlinks
        if(column_letter == 'A'):
            for cell in col[1:]:
                cell.hyperlink = cell.value
                cell.style = "Hyperlink"


def BuildXlsxFile(name, data_findings, findings_column_colors, data_analytics):
    wb, ws_analytics, ws_findings = Setup()

    for row in data_analytics:
        ws_analytics.append(row)

    SetColumnSize(ws_analytics)

    for row in data_findings:
        ws_findings.append(row)

    SetFindingsHyperlinks(ws_findings)
    SetColumnColors(ws_findings, findings_column_colors)
    SetColumnSize(ws_findings)

    SaveFile(wb, name)